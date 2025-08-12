import openpyxl


class Excel_utilities_class:

    @staticmethod
    def getRowCount (file, sheet_name):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.max_row

    @staticmethod
    def read_data( file, sheet_name, rownum, columnno):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        return sheet.cell(row=rownum, column=columnno).value

    @staticmethod
    def write_data( file, sheet_name, rownum, columnno, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        sheet.cell(row=rownum, column=columnno).value = data
        workbook.save(file)
        workbook.close()